"""
Results caching and export.
"""

__author__ = "Jacob Bumgarner <jrbumgarner@mix.wvu.edu>"
__license__ = "GPLv3 - GNU General Pulic License v3 (see LICENSE)"
__copyright__ = "Copyright 2022 by Jacob Bumgarner"
__webpage__ = "https://jacobbumgarner.github.io/VesselVio/"
__download__ = "https://jacobbumgarner.github.io/VesselVio/Downloads"


import csv
import os
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from pyexcelerate import Alignment, Font, Style, Workbook

from library import helpers


###########################
### Results Header Text ###
###########################
## Main results header
def load_headers():
    results_topper = [""] * 76
    results_topper[2] = "Main Results"
    results_topper[15] = "Number of Segments per Radius Bin"
    results_topper[36] = "Mean Length of Segments per Radius Bin"
    results_topper[57] = "Mean Segment Tortuosity per Radius Bin"
    results_header = [
        "File Name",
        "ROI Name",
        "ROI_Volume",
        "Volume",
        "Network Length",
        "Surface Area",
        "Branchpoints",
        "Endpoints",
        "Number of Segments",
        "Segment Partitioning",
        "Mean Segment Radius",
        "Mean Segment Length",
        "Mean Segment Tortuosity",
        "Mean Segment Volume",
        "Mean Segment Surface Area",
    ]
    bins = []
    for i in range(3):
        for i in range(20):
            bin = str(i) + " - " + str(i + 1)
            results_header.append(bin)
        results_header.append("20+")
    results_header = [results_topper, results_header]

    ## Segment results header
    segment_results_header = [
        "Segment ID",
        "Volume",
        "Length",
        "Surface Area",
        "Tortuosity",
        "Mean Radius",
        "Max Radius",
        "Min Radius",
        "Radius Std. Dev.",
    ]
    return results_header, segment_results_header


def create_results_file(results_file, data):
    wb = Workbook()
    ws = wb.new_sheet("Main Results", data=data)
    row_style = Style(
        font=Font(bold=True),
        alignment=Alignment(wrap_text=True, horizontal="center", vertical="center"),
        size=60,
    )
    # col_style = Style(font=Font(bold=True), alignment=Alignment(wrap_text=True, horizontal='fill', vertical='bottom'))
    col_style = Style(font=Font(bold=True))
    ws.set_col_style((3, 7, 9, 10, 15, 16, 37, 58), Style(size=12))
    ws.set_row_style((1, 2), row_style)
    ws.set_col_style((1, 2), col_style)
    wb.save(results_file)


def read_ws(ws):
    results = []
    for row in ws.iter_rows():
        results.append([cell.value for cell in row])
    return results


def write_results(results_folder, image_dimensions=3, verbose=False):
    if verbose:
        print("Exporting results...", end="\r")
    if os.path.exists(results_folder) == False:
        os.mkdir(results_folder)

    results_header, segment_results_header = load_headers()

    if image_dimensions == 2:
        results_header[1][3] = "Percent Area Fraction (PAF) %"
        results_header[1][13] = "Mean Segment PAF %"
        segment_results_header[1] = "Segment Percent Area Fraction %"
        results_file = os.path.join(
            results_folder, "VesselVio 2D Dasataset Analysis Results.xlsx"
        )
    else:
        results_file = os.path.join(
            results_folder, "VesselVio 3D Dataset Analysis Results.xlsx"
        )

    results_file = helpers.std_path(results_file)
    if not os.path.exists(results_file):
        results = results_header + read_cache_results()
        create_results_file(results_file, results)

    else:
        # Kind of silly to do it like this, but oh well.
        # This keeps the formatting nice
        # xlsx writing is already slow enough as it is
        wb = load_workbook(results_file, read_only=True)
        ws = wb["Main Results"]
        results = read_ws(ws) + read_cache_results()
        create_results_file(results_file, results)
        wb.close()

    # Delete the results cache file
    delete_results_cache()
    return


def write_seg_results(seg_results, results_folder, filename, ROI_Name):
    _, segment_results_header = load_headers()

    # Make sure the folder exists
    if not os.path.exists(results_folder):
        os.mkdir(results_folder)
    segments_folder = os.path.join(results_folder, "Segment Results")
    if not os.path.exists(segments_folder):
        os.mkdir(segments_folder)

    # Add the ROI name if it exists
    if ROI_Name != "None":
        file = os.path.join(segments_folder, filename + "_" + ROI_Name + ".csv")
    else:
        file = os.path.join(segments_folder, filename + ".csv")

    # Save the info
    with open(file, "w") as f:
        writer = csv.writer(f)
        writer.writerow(["Filename:", filename, "ROI Name:", ROI_Name])
        writer.writerow(segment_results_header)
        writer.writerows(seg_results)
    return


################################
### Results Cache Processing ###
################################
# Store the result in the cache csv file.
def cache_result(result):
    results_cache = get_cache_path()
    if not os.path.exists(results_cache):
        with open(results_cache, "w") as f:
            writer = csv.writer(f)
            writer.writerow(result)
    else:
        with open(results_cache, "a") as f:
            writer = csv.writer(f)
            writer.writerow(result)
    return


# Read the results stored in the cache csv.
def read_cache_results():
    results_cache = get_cache_path()
    results = []
    if os.path.exists(results_cache):
        with open(results_cache, "r") as f:
            reader = csv.reader(f)
            for result in reader:
                for i in range(len(result)):
                    try:
                        result[i] = Decimal(result[i]).quantize(Decimal("1.000000"))
                    except InvalidOperation:
                        continue
                results.append(result)
    return results


def get_cache_path():
    wd = helpers.get_cwd()
    results_cache = helpers.std_path(
        os.path.join(wd, "library", "cache", "results_cache.csv")
    )
    return results_cache


# Delete cache file after successfully exporting the results
def delete_results_cache():
    results_cache = get_cache_path()
    if os.path.exists(results_cache):
        os.remove(results_cache)
    return
