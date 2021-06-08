# -*- coding: utf-8 -*-

# Load neessary packages.
from os import listdir, path, mkdir

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

import numpy as np
import time

from geomdl import BSpline, knotvector, operations

import igraph as ig

from SimpleITK import GetImageFromArray, WriteImage


results_topper = ['Main Results', ' Number of Segments per Radius Bin', 'Mean Length of Segments per Radius Bin', 'Mean Segment Tortuosity per Radius Bin']
results_header = ['File Name', "Volume/Area", 'Skeleton Length', 'Branchpoints', 'Endpoints', 
            'Number of Segments', 'Mean Segment Length', 'Mean Segment Radius', 'Mean Segment Tortuosity', 'Segment Partitioning']
bins = []
for i in range(20):
    a = str(i)
    b = str(i + 1)
    binned = a + ' - ' + b
    bins.append(binned)
bins.append('20+')
for i in range(3):
    results_header.extend(bins)

segments_headers = ['Segment Number', 'Mean Radius', 'Length', 'Tortuosity']


# Create a list of the coords and radii of the vertices of extracted 1-vertex segments.
# (Can either be 2- or 3-vertices long)
def small_segs(g, gsegs, segment, segment_ids, Visualization=False, pruning=False):
    # Prepare necessary lists
    point_list = []
    coords_list = []
    radii_list = []
    
    segment_length = 0
    vertex = segment[0]
    
    # Find neighbors of the single vertex segment, then extract the coords and radii of the ordered points.
    # There will either be one or two neighbors
    true_vert = segment_ids[vertex].index
    
    neighbors = g.neighbors(true_vert)
    edges = g.incident(true_vert)
    
    # Populate lists with vertex info. 3rd vertex may not exist.
    for i in range(3):
        if i == 0:
            v = neighbors[0]
        if i == 1:
            v = true_vert
        if i == 2:
            if len(neighbors) > 1:
                v = neighbors[1]
        vi = g.vs[v]
        point_list.append(vi)
        coords_list.append(vi['v_coords'])
        radii_list.append(vi['v_radius'])
        
        
    # EDT Between points, stored as 'edge_length', see Graph_Processing
    for edge in edges:
        segment_length += g.es[edge]['edge_length']
    if pruning: 
        return segment_length
        
    # Send back avg_radius
    avg_radius = np.mean(radii_list)
    
    # Calculate the tortuosity. Divide segment length by cord length of endpoints.
    coords_list = np.array(coords_list)
    delta = coords_list[-1] - coords_list[0]
    cord_length = np.linalg.norm(delta)

    # Needed to deal with loops. Unlikely in these small segments, but just in case...
    if cord_length > 0:
        tortuosity = segment_length / cord_length
    else:
        tortuosity = 0
    
    # Visualization checkpoint.
    if Visualization == False:
        return segment_length, avg_radius, tortuosity
    
    else:
        end_degrad = []
        e1 = point_list[0]
        e2 = point_list[-1]
        e1deg = g.degree(e1)
        e1rad = e1['v_radius']
        e1info = [e1deg, e1rad]
        end_degrad.append(e1info)
        e2deg = g.degree(e2)
        e2rad = e2['v_radius']
        e2info = [e2deg, e2rad]
        end_degrad.append(e2info)
        
        return coords_list, radii_list, segment_length, avg_radius, tortuosity, end_degrad



# Create a list of the coords and radii of the vertices of extracted segments.
def large_segs(g, gsegs, segment, segment_ids, Visualization=False, pruning=False):
    # Prepare necessary lists
    point_list = []
    radii_list = []
    endpoints = []
    
    # First find the endpoints of our segment.
    for vertex in segment:
        if gsegs.degree(vertex) == 1:
            endpoints.append(vertex)
            if len(endpoints) == 2:
                break
    

    if len(endpoints) == 2:
        # Find the ordered path of vertices between each endpoint.
        # The indices of this path will be relative
        path = gsegs.get_shortest_paths(endpoints[0], to=endpoints[1], output='vpath')[0]

        # Add true indices of our segment path to the point_list.
        for point in path:
            point_list.append(segment_ids[point].index)
        
        # Extend the point_list by any neighbors on either end of the segment.
        e1rn = segment_ids[gsegs.neighbors(endpoints[0])[0]].index
        e2rn = segment_ids[gsegs.neighbors(endpoints[1])[0]].index
        
        e1ns = g.neighbors(segment_ids[endpoints[0]].index)
        e2ns = g.neighbors(segment_ids[endpoints[1]].index)
        
        # Add first potential extension to beginning of list.
        for neighbor in e1ns:
            if neighbor != e1rn:
                point_list.insert(0, neighbor)
            
        # Add second potential extension to end of list.
        for neighbor in e2ns:
            if neighbor != e2rn:
                point_list.append(neighbor)
                
        # Extract the radii and coordinates of our point list. 
        radii_list = (g.vs[point_list]['v_radius'])
        point_coords = (g.vs[point_list]['v_coords'])
    
    # Loops in the vasculature...
    elif len(endpoints) != 2:
        try:
            circ_ordered = []
        
            # Choose a random first point.
            first = segment[4] # 0 kept crashing my runs and leading to WEIRD issues. No idea.
            circ_ordered.append(first)
            
            # Choose random neighbor of that point.
            n1 = gsegs.neighbors(first)[0]
            last = n1
            circ_ordered.append(n1)
            
            # Find neighbor of n1 that isn't first.
            n1ns = gsegs.neighbors(n1)
            for n in n1ns:
                if n != first:
                    n1 = n
                    circ_ordered.append(n)
                        
            # Now loop through the points until we find the first point.
            looped = False
            i = 0
            size = len(segment)
            while looped == False:
                if i > size: 
                    looped = True
                    break
                ns = gsegs.neighbors(n1)
                for n in ns:
                    if looped == True: # Just to get out of the ns iter.
                        break
                    if n == first:
                        circ_ordered.append(n)
                        looped = True
                        break
                    elif n != last:
                        last = n1
                        n1 = n
                        circ_ordered.append(n)
                    elif n in circ_ordered:
                        looped = True
                        break
                i += 1
                
            # Convert back into graph units.
            for point in circ_ordered:
                point_list.append(segment_ids[point].index)

            # Get our radii and coords
            radii_list = (g.vs[point_list]['v_radius'])
            point_coords = (g.vs[point_list]['v_coords'])
        
        # If this doesn't work, just return an approximation of the loop size. Had some extreme oddities and errors when catching this.
        except:
            segment_length = size
            if pruning:
                return segment_length
            avg_radius = np.mean(radii_list)
            tortuosity = 0
            if Visualization == False:
                return segment_length, avg_radius, tortuosity
            else:
                end_degrad = [[2, 1],[2,1]]
                for i in range(3):
                    point_list.append(segment_ids[segment[i]].index)
                coords_list = g.vs[point_list]['v_coords']
                return coords_list, radii_list, segment_length, avg_radius, tortuosity, end_degrad #F
                
                
    ## BSpline
    # Find basis-spline (BSpline) of long segment paths to smooth jaggedness of skeleton paths.
    # The mathematics of BSplines can be examined further below:
        # http://www.independent-software.com/determining-coordinates-on-a-html-canvas-bezier-curve.html
        # https://web.mit.edu/hyperbook/Patrikalakis-Maekawa-Cho/node17.html
        # http://learnwebgl.brown37.net/07_cameras/points_along_a_path.html
        
    # Here I have used NURBs (Bingol and Krishnamurthy, 2019)
    point_coords = np.array(point_coords) # Needed to clear array status of elements.
    point_coords = point_coords.tolist() # Geomdl doesn't cooperate with numpy.

    num_verts = len(point_coords)
    
    # There should be no segments in this function smaller than 3-verts in size.
    # Because we removed small, isolated segments from our graph, and filtered small segments to small_segs()
    
    # Set appropriate degree of our BSpline
    if num_verts > 4:
        spline_degree = 4
    else:
        spline_degree = num_verts - 1
    
    # Find the segment length based on our cubic BSpline.
    curve = BSpline.Curve()
    curve.degree = spline_degree
    curve.ctrlpts = point_coords
    curve.knotvector = knotvector.generate(curve.degree, curve.ctrlpts_size)
    
    
    # The optimal number of interpolated segment points for visualization was determined emperically as a trade-off value between ground-truth length and computational costs.
    if num_verts < 5:
        correction = 1
    elif num_verts < 10:
        correction = 2
    elif num_verts < 20:
        correction = 3
    elif num_verts < 40:
        correction = 4
    elif num_verts < 80:
        correction = 5
    elif num_verts < 160:
        correction = 8
    else: 
        correction = 12
    delta = 1 / (num_verts / correction)
    curve.delta = delta # n_points = 1/delta
    segment_length = operations.length_curve(curve)
    
    if pruning:
        return segment_length
    
    if Visualization:
        coords_list = curve.evalpts
        coords_list = np.array(coords_list)    

    # coords_list = np.array(point_coords)
        
    # Calculate the average radius
    avg_radius = np.mean(radii_list)

    # Calculate the tortuosity. Divide segment length by cord length of endpoints.
    point_coords = np.array(point_coords)
    delta = point_coords[-1] - point_coords[0]
    cord_length = np.linalg.norm(delta)
    
    # This is needed to account for loops.
    if cord_length > 0:
        tortuosity = segment_length / cord_length
    else:
        tortuosity = 0
        
    if Visualization == False:
        return segment_length, avg_radius, tortuosity
    
    else:
        end_degrad = []
        e1 = g.vs[point_list[0]]
        e2 = g.vs[point_list[-1]]
        e1deg = g.degree(e1)
        e1rad = e1['v_radius']
        e1info = [e1deg, e1rad]
        end_degrad.append(e1info)
        e2deg = g.degree(e2)
        e2rad = e2['v_radius']
        e2info = [e2deg, e2rad]
        end_degrad.append(e2info)
        
        return coords_list, radii_list, segment_length, avg_radius, tortuosity, end_degrad


# Isolates segments from dataset, then ships them off for feature extraction.
def segment_prep(g, verbose = False):
    t1 = time.perf_counter()
    
    ## Find segments of vertices between branchpoints & endpoints.
    # Isolate all vertices that aren't branchpoints.
    segment_ids = g.vs.select(_degree_lt = 3)
    
    # Create a subgraph of these vertices with their edges to find segments between branchpoints.
    gsegs = g.subgraph(segment_ids)
    
    # Find the individual segments
    segments = gsegs.clusters()
    
    segment_count = 0
    
    segments_dict = []  
    avg_radius_list = []
    segment_length_list = []
    tortuosity_list = []
    
    total_length = 0
    
    # Iterate through each segment.
    # Finds the the average radius, length, and tortuosity (arc-cord ratio) of segments extended to branchpoint(s).
    for segment in segments:     
        segment_count += 1   
        seg_size = len(segment)
        
        ## Segment preparation
        # Prepare point list for single-vertex segments
        if seg_size == 1:
            segment_length, avg_radius, tortuosity = small_segs(g, gsegs, segment, segment_ids)
        
        # Prepare point list for two-vertex segments.
        elif seg_size > 1:
            segment_length, avg_radius, tortuosity = large_segs(g, gsegs, segment, segment_ids)
        
        # Add segment length to total length of dataset
        total_length += segment_length
        
        ## We can store all of these in a segment dictionary, and append it to our main dict.
        avg_radius_list.append(avg_radius)
        segment_length_list.append(segment_length)
        tortuosity_list.append(tortuosity)
                
    t2 = time.perf_counter()
    
    ## Find segments between branchpoints
    bif_ids = g.vs.select(_degree_gt = 2)
    gbifs = g.subgraph(bif_ids)
    
    for e in gbifs.es():
        segment_count += 1
        total_length += e['edge_length']
        
        v1 = gbifs.vs[e.source]
        v2 = gbifs.vs[e.target]
        avg_radius = (v1['v_radius'] + v2['v_radius']) / 2
        avg_radius_list.append(avg_radius)
        segment_length_list.append(e['edge_length'])
        tortuosity_list.append(1)
    
            
    if verbose:
        print(f"Segments analyzed in {t2-t1:0.2f} s")
    
    avg_radius_list = np.array(avg_radius_list)
    segment_length_list = np.array(segment_length_list)
    tortuosity_list = np.array(tortuosity_list)
    
    segments_dict = {'avg_radius': avg_radius_list,
                'segment_length': segment_length_list,
                'tortuosity': tortuosity_list}
    
    
    return segment_count, total_length, segments_dict

        
# Main window for extracting the features from our dataset.
def features(g, volume, resolution, filename, save_seg_results=False, save_labeled=False):
    # Find our filename.
    name = path.splitext(path.basename(filename))[0]    
    
    # Calculate our volume, branchpoints, endpoints
    total_volume = np.sum(volume) * resolution
    bs = g.vs.select(_degree_gt = 2)
    branchpoints = len(bs)
    es = g.vs.select(_degree = 1)
    endpoints = len(es)
        
    # Ship our graph away to extract segment count, segment radius, skeletal length, and tortuosity information.
    segment_count, total_length, segments_dict = segment_prep(g)    
    
    # Pull radii and lengths. Need these for histograms and averages.
    radii = segments_dict.get('avg_radius') * resolution
    lengths = segments_dict.get('segment_length') * resolution
    tortuosities = segments_dict.get('tortuosity')
    
    # Start preparing and loading our results for xlsx export.
    total_length *= resolution
    segment_partitioning = segment_count / total_length
    
    avg_radius = np.mean(radii)
    avg_length = np.mean(lengths)
    avg_tortuosity = np.mean(tortuosities)
    
    
    # Add current results to our results list.
    results = [name, total_volume, total_length, branchpoints, endpoints, segment_count, avg_length, avg_radius, avg_tortuosity, segment_partitioning]
    
    # Find histogram of radii distributions    
    bins = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 500]
    distribution = np.histogram(radii, bins)[0]
    
    # Add these histogram values to our results.
    for r_bin in distribution:
        if r_bin > 0:
            results.append(r_bin)
        else:
            results.append(0)
        
    # Find mean tortuosity & lengths of vessels in bins. First find tortuosity sums for vessels.
    tort_bins = np.zeros(21)
    tort_counts = np.zeros(21)
    len_bins = np.zeros(21)
    len_counts = np.zeros(21)
    
    for radius in radii:
        for i in range(21):
            in_bin = False
            if i < 20:
                if radius < i + 1:
                    in_bin = True
            elif i == 20:
                if radius >= 20:
                    in_bin = True
            
            if in_bin:
                tort_bins[i] += tortuosities[i]
                tort_counts[i] += 1
                len_bins[i] += lengths[i]
                len_counts[i] += 1
                break
            else:
                continue
            
    # Then find the means of the sums and append the results to our results list.
    for i in range(21):
        if len_counts[i] > 0:
            len_bins[i] = len_bins[i] / len_counts[i]
            results.append(len_bins[i])
        else:
            results.append('')
            
    for i in range(21):
        if tort_counts[i] > 0:
            tort_bins[i] = tort_bins[i] / tort_counts[i]
            results.append(tort_bins[i])
        else:
            results.append('')
                 
     
    ### Save labelled skeleton ###
    ## Only implemented for command-line interaction. 
    if save_labeled:
        if path.exists("./Labeled") == False:
            mkdir("./Labeled")
        test_array = np.zeros(volume.shape, dtype='uint8')
        for v in g.vs():
            z,y,x = v["v_coords"]
            test_array[z, y, x] = 130
        for b in bs:
            z, y, x = b["v_coords"]
            test_array[z, y, x] = 255
        for e in es:
            z, y, x = e["v_coords"]
            test_array[z, y, x] -= 10
    
    
        image = GetImageFromArray(test_array)
        name = "Labeled/" + name + '.nii'
        WriteImage(image, name)
     
    if not save_seg_results: 
        return results
    else:
        seg_results = [segments_dict['avg_radius'].tolist(), segments_dict['segment_length'].tolist(), segments_dict['tortuosity'].tolist()]
        headers = ['Average Radius', 'Average Length', 'Tortuosity']
        for i in range(3):
            header = ''
            if i == 0:
                header = name
            seg_results[i].insert(0, segments_headers[i])
            seg_results[i].insert(0, header)
        return results, seg_results
             

# Write the results to an exported CSV file.
def write_results(results, results_folder, seg_results=None):
    # Short function to format our header cells.
    def cell_format(cell):
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)           
        column = get_column_letter(cell.column)
        ws.column_dimensions[column].auto_size = True
        
    # Check to make sure our directory exists. If not, create it.
    if path.exists(results_folder) == False:
        mkdir(results_folder)
        
    dir_files = listdir(results_folder)
    file_path = path.join(results_folder, "VesselVio Analysis Results.xlsx")
    
    # Check to see if our default save file is present. If not, make new one and append headers.
    if "VesselVio Analysis Results.xlsx" not in dir_files:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Main Results'

        # Add header information
        ws['B1'] = results_topper[0]       
        ws['K1'] = results_topper[1]      
        ws['AF1'] = results_topper[2]
        ws['BA1'] = results_topper[3] 
        ws.append(results_header)
        
        # Append results
        for i in range(len(results)):
            ws.append(results[i])
        
        for header_cell in ws['1:1']:
            cell_format(header_cell)
        
        for header_cell in ws['2:2']:
            cell_format(header_cell)
        
        ws.row_dimensions[1].bestFit = True
        
        first_cell = ws['A3']
        cell_format(first_cell)        
        
        for column_cell in ws['A:A']:
            cell_format(column_cell)       
        
        wb.save(file_path)
        wb.close()
    
    # If file is present, append data below lowest row.
    else:
        wb = load_workbook(file_path)
        ws = wb.active
        for i in range(len(results)):
            ws.append(results[i])
        
        for column_cell in ws['A:A']:
            cell_format(column_cell)       
        
        wb.save(file_path)
        wb.close() 
        
    # If the segments dict is sent, write it to the segments dict sheet.
    if seg_results:
        wb = load_workbook(file_path)
        sheet_names = wb.sheetnames
        
        # Create segments sheet if it doesn't exist.
        if 'Segment Results' not in sheet_names:
            wb.create_sheet('Segment Results')
        
        ws = wb['Segment Results']
            
        # Find out how many datasets there are.
        num_results = len(seg_results)
        added = 0
        for i in range(num_results):
            num_segs = len(seg_results[i][0])
            starting_column = ws.max_column
            if starting_column != 1:
                starting_column += 1
                added = 1
                
            for j in range(3):
                column = starting_column + j + added
                k = 0
                
                # Iterate over the column cells 
                for col in ws.iter_cols(min_row=2, max_row=num_segs+1, min_col=column, max_col=column):
                    for cell in col:
                        cell.value = seg_results[i][j][k]
                        k += 1
                
            
        wb.save(file_path)
        wb.close()
        