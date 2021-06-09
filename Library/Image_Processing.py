# -*- coding: utf-8 -*-


import sys, os
import numpy as np
import time

# Image processing
from pathlib import Path
from PIL import Image
from SimpleITK import ImageFileReader, GetArrayFromImage, GetImageFromArray, WriteImage

# Result Export
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

##########################
#### Array Processing ####
##########################

## Returns a true binary (0,1) array from an image file when given the file name and directory.
def getArray(file, verbose=False):
    t1 = time.perf_counter()
    
    reader = ImageFileReader()
    reader.SetFileName(file)
    try:
        image = reader.Execute() 
    except:
        return None
    data = GetArrayFromImage(image)
    
    data = (data > 0).astype(np.uint8)
    data = np.ascontiguousarray(data, dtype=np.uint8)
    data = np.pad(data, 1)
    
    if data.ndim not in [2,3]:
        return 'Image not compatible.'
    
    if verbose:
        print (f"Array loaded in {time.perf_counter() - t1: 0.4f} s.")
    
    return data
   
# Get the sum of an ROI volume and correct it with the resolution.   
def getROIVolume(file, resolution):
    reader = ImageFileReader()
    reader.SetFileName(file)
    try:
        image = reader.Execute() 
    except:
        return None
    data = GetArrayFromImage(image)
    
    binary = data > 0
    dataset = binary * 1 * resolution
    result = np.sum(dataset)
    
    return result
    
    
##########################
#### Image Processing ####
##########################
    
## Given folder name and arguments from main, loads sequence of 
## images, binarizes them, and exports them as a single image stack.
def sequencer(file_dir, output_dir, sub_dir, old_ext='.bmp', new_ext='.nii'):
    
    # Use directory name for our filename, make temp file list.
    filename = os.path.basename(file_dir)  
    temp = []                       
        
    # Get Path of dir/subfolder so we can use glob for the image names.
    if sub_dir:
        file_dir = Path(file_dir, sub_dir)
    else:
        file_dir = Path(file_dir)

    # Iterate through our folder to find files with our desired .ext
    images = sorted([str(fn) for fn in file_dir.glob('*'+old_ext)])
    
    if not images:
        return
    
    # Opens each file in series and appends it to the 'temp' list.
    for fn in images:
        img = np.asarray(Image.open(fn)) # Had to use PIL here rather than SITK
        # img = img
        temp.append(img)
    
    # Loads tmp list as an 8-bit array into img, which is then converted and written into an image.
    temp = np.asarray(temp, dtype='uint8')    
    # Just ensures that we end up with a 255 binary.
    binary = temp > 0
    dataset = binary * 255

    dataset = dataset.astype(np.uint8)
    
    # Find the path for our new filename.
    filename = filename + new_ext
    filepath = os.path.join(output_dir, filename)
    
    image = GetImageFromArray(dataset)
    WriteImage(image, filepath)
    return
    
    
# Iterate through the directory input and send the files off for volume calculation.
def ROIV_input(directory, results_dir, ext, resolution, verbose=False):
    results = []
    for file in os.listdir(directory):
        if file.endswith(ext):
            filename = os.path.splitext(file)[0]
            file_path = os.path.join(directory, file)
            volume = getROIVolume(file_path, resolution)
            
            result = [filename, volume]
            results.append(result)
    
    # Save the results to our output path.
    def cell_format(cell): # Reusing here because of laziness
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrapText=True)           
        column = get_column_letter(cell.column)
        ws.column_dimensions[column].auto_size = True
    
    if os.path.exists(results_dir) == False:
        os.mkdir(results_dir)
    
    dir_files = os.listdir(results_dir)
    file_path = os.path.join(results_dir, "ROI Volumes.xlsx")
    if file_path not in results_dir:
        wb = Workbook()
        ws = wb.active
        ws.title = 'ROI Volumes'
        header = ['File', 'Volume']
        ws.append(header)
        cell2 = ws['B1']
        cell_format(cell2)
        
        wb.save(file_path)
        wb.close()
    
    wb = load_workbook(file_path)
    ws = wb.active
    for i in range(len(results)):
        ws.append(results[i])
    
    for column_cell in ws['A:A']:
        cell_format(column_cell)
    
    wb.save(file_path)
    wb.close()
        
    return
    
# Iterate through the directory input and send the subfolder off to be processed.
def sequencer_input(dir_path, output_dir, sub_dir=None, old_ext='.bmp', new_ext='.nii', verbose=False):
    if os.path.exists(dir_path) == False:
        os.mkdir(dir_path)
    if os.path.exists(output_dir) == False:
        os.mkdir(output_dir)
    for folder in os.listdir(dir_path):
        folder_path = os.path.join(dir_path, folder)
        if os.path.isdir(folder_path):
            if verbose:
                print ("Processing:", folder_path, "...", end='')
                t1 = time.perf_counter()
            sequencer(folder_path, output_dir, sub_dir, old_ext, new_ext)
            if verbose:
                print (f"√\n     Image series created in {time.perf_counter() - t1:0.2f} seconds.")
    
    return    
    
if __name__ == "__main__":
    # sequencer_input(directory, output, sub_dir, verbose=True)
    directory = '/Volumes/Bumgarner/NII Folder'
    results_directory = '/Users/jacobbumgarner/Desktop/VesselVio'
    sub_directory = 'Volume'
    old_ext = '.bmp'
    new_ext = '.nii'
    ext = '.nii'
    resolution = 2.7
    
    sequencer_input(directory, results_directory, sub_directory, 
                    old_ext, new_ext, verbose=True)
    # ROIV_input(directory, results_directory, ext, resolution, verbose=True)
